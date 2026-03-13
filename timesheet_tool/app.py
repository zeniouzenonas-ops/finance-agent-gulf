import io
import re
import os
from datetime import datetime

import pdfplumber
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import streamlit as st

# ── Day-type sets ───────────────────────────────────────────────────────────
CHARGEABLE  = {"W", "HD", "OFF", "WH"}
TRAVEL      = {"T"}

# ── Colours ─────────────────────────────────────────────────────────────────
C_DARK_BLUE = "1F4E79"
C_GREEN     = "C6EFCE"
C_RED       = "FFCCCC"
C_AMBER     = "FFEB9C"


# ═══════════════════════════════════════════════════════════════════════════
# PDF PARSING
# ═══════════════════════════════════════════════════════════════════════════

def extract_header(text: str) -> dict:
    info = {"name": "", "project": "", "month": "", "year": "", "bu": ""}

    m = re.search(r"\*?NAME:\s*(.+?)(?:\s{2,}|\s+YEAR:)", text)
    if m: info["name"] = m.group(1).strip()

    m = re.search(r"\*?PROJECT:\s*(.+?)(?:\s{2,}|\s+\*?MONTH:)", text)
    if m: info["project"] = m.group(1).strip()

    m = re.search(r"\*?MONTH:\s*(\w+)", text)
    if m: info["month"] = m.group(1).strip()

    m = re.search(r"YEAR:\s*(\d{4})", text)
    if m: info["year"] = m.group(1).strip()

    m = re.search(r"(BU-\d+)", text)
    if m: info["bu"] = m.group(1).strip()

    return info


def safe_col(row, idx):
    if idx < len(row) and row[idx] is not None:
        return str(row[idx]).strip()
    return ""


def extract_daily_rows(table) -> list:
    rows = []
    for row in (table or []):
        if not row or not row[0]:
            continue
        date_str = str(row[0]).strip()
        if not re.match(r"\d{1,2}/\d{1,2}/\d{4}", date_str):
            continue

        day_shift  = safe_col(row, 2)
        night_shift= safe_col(row, 3)
        shift_type = day_shift if day_shift else night_shift
        st_upper   = shift_type.upper()

        if   st_upper in CHARGEABLE: charged = "YES"
        elif st_upper in TRAVEL:     charged = "CHECK"
        else:                        charged = "NO"

        rows.append({
            "Date":        date_str,
            "Location":    safe_col(row, 1),
            "Shift Type":  shift_type,
            "Start":       safe_col(row, 4),
            "End":         safe_col(row, 5),
            "Break":       safe_col(row, 6),
            "Total Hours": safe_col(row, 7),
            "Remarks":     safe_col(row, 8),
            "Charged":     charged,
        })
    return rows


def parse_pdf(pdf_bytes: bytes) -> list:
    timesheets = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            try:
                text   = page.extract_text() or ""
                header = extract_header(text)
                if not header["name"]:
                    continue
                daily_data = extract_daily_rows(page.extract_table())
                timesheets.append({**header, "page": page_num, "daily_data": daily_data})
            except Exception as exc:
                st.warning(f"Page {page_num}: could not parse ({exc})")
    return timesheets


# ═══════════════════════════════════════════════════════════════════════════
# RATES
# ═══════════════════════════════════════════════════════════════════════════

def load_rates(excel_bytes: bytes) -> dict:
    df = pd.read_excel(io.BytesIO(excel_bytes))
    rates = {}
    for _, row in df.iterrows():
        name    = str(row.iloc[0]).strip()
        emp_num = str(row.iloc[1]).strip()
        try:    rem  = float(row.iloc[2])
        except: rem  = 0.0
        try:    cost = float(row.iloc[3])
        except: cost = 0.0
        rates[name.lower()] = {
            "employee_number":   emp_num,
            "remuneration_rate": rem,
            "cost_rate":         cost,
        }
    return rates


# ═══════════════════════════════════════════════════════════════════════════
# REPORT
# ═══════════════════════════════════════════════════════════════════════════

def _style_header(ws, num_cols):
    fill = PatternFill("solid", fgColor=C_DARK_BLUE)
    font = Font(bold=True, color="FFFFFF")
    for col in range(1, num_cols + 1):
        c = ws.cell(row=1, column=col)
        c.fill = fill; c.font = font
        c.alignment = Alignment(horizontal="center", vertical="center")

def _autofit(ws, max_w=35):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        w = max((len(str(c.value)) for c in col if c.value), default=8)
        ws.column_dimensions[letter].width = min(w + 2, max_w)


def generate_report(timesheets: list, rates: dict) -> bytes:
    wb = Workbook()

    # ── Detailed sheet ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Detailed Report"
    DETAIL_COLS = [
        "Employee Name", "Employee #", "Project", "BU", "Month", "Year",
        "Date", "Location", "Shift Type", "Start", "End",
        "Total Hours", "Remarks", "Charged",
        "Remuneration Rate", "Cost Rate",
        "Remuneration Amount", "Cost Amount",
    ]
    ws.append(DETAIL_COLS)
    _style_header(ws, len(DETAIL_COLS))

    yes_fill   = PatternFill("solid", fgColor=C_GREEN)
    no_fill    = PatternFill("solid", fgColor=C_RED)
    check_fill = PatternFill("solid", fgColor=C_AMBER)

    row_num = 2
    for ts in timesheets:
        name      = ts["name"]
        r         = rates.get(name.lower(), {})
        emp_num   = r.get("employee_number",   "NOT FOUND")
        rem_rate  = r.get("remuneration_rate", 0)
        cost_rate = r.get("cost_rate",         0)

        for day in ts["daily_data"]:
            charged = day["Charged"]
            rem_amt  = rem_rate  if charged == "YES" else ("CHECK" if charged == "CHECK" else 0)
            cost_amt = cost_rate if charged == "YES" else ("CHECK" if charged == "CHECK" else 0)

            ws.append([
                name, emp_num, ts["project"], ts["bu"], ts["month"], ts["year"],
                day["Date"], day["Location"], day["Shift Type"],
                day["Start"], day["End"], day["Total Hours"], day["Remarks"],
                charged, rem_rate, cost_rate, rem_amt, cost_amt,
            ])
            c = ws.cell(row=row_num, column=14)
            c.fill = yes_fill if charged == "YES" else (check_fill if charged == "CHECK" else no_fill)
            row_num += 1

    _autofit(ws)
    ws.freeze_panes = "A2"

    # ── Summary sheet ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    SUM_COLS = [
        "Employee Name", "Employee #", "Project", "BU", "Month", "Year",
        "W Days", "HD Days", "Off Days", "WH Days",
        "T Days (manual check)", "Non-Chargeable Days", "Total Chargeable Days",
        "Remuneration Rate", "Total Remuneration",
        "Cost Rate", "Total Cost",
    ]
    ws2.append(SUM_COLS)
    _style_header(ws2, len(SUM_COLS))

    for ts in timesheets:
        name      = ts["name"]
        r         = rates.get(name.lower(), {})
        emp_num   = r.get("employee_number",   "NOT FOUND")
        rem_rate  = r.get("remuneration_rate", 0)
        cost_rate = r.get("cost_rate",         0)

        cnt = {"W": 0, "HD": 0, "OFF": 0, "WH": 0, "T": 0, "OTHER": 0}
        chargeable_days = 0
        for day in ts["daily_data"]:
            st = day["Shift Type"].upper()
            if   st == "W":   cnt["W"]    += 1
            elif st == "HD":  cnt["HD"]   += 1
            elif st == "OFF": cnt["OFF"]  += 1
            elif st == "WH":  cnt["WH"]   += 1
            elif st == "T":   cnt["T"]    += 1
            else:             cnt["OTHER"]+= 1
            if day["Charged"] == "YES":
                chargeable_days += 1

        ws2.append([
            name, emp_num, ts["project"], ts["bu"], ts["month"], ts["year"],
            cnt["W"], cnt["HD"], cnt["OFF"], cnt["WH"],
            cnt["T"], cnt["OTHER"], chargeable_days,
            rem_rate, chargeable_days * rem_rate,
            cost_rate, chargeable_days * cost_rate,
        ])

    _autofit(ws2)
    ws2.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════
# STREAMLIT UI
# ═══════════════════════════════════════════════════════════════════════════

st.set_page_config(page_title="Timesheet Processor", page_icon="📋", layout="centered")

st.markdown(
    """
    <h1 style='color:#1F4E79;'>📋 Timesheet Processor</h1>
    <p style='color:#555;'>Upload the combined PDF timesheets and your employee rates file to generate a billing report.</p>
    """,
    unsafe_allow_html=True,
)

st.divider()

col1, col2 = st.columns(2)
with col1:
    pdf_file = st.file_uploader("📄 Combined PDF Timesheets", type=["pdf"])
with col2:
    excel_file = st.file_uploader("📊 Employee Rates (Excel)", type=["xlsx", "xls"])

st.divider()

if st.button("⚡ Generate Report", type="primary", use_container_width=True):
    if not pdf_file or not excel_file:
        st.error("Please upload both files before generating the report.")
    else:
        with st.spinner("Reading PDF timesheets…"):
            timesheets = parse_pdf(pdf_file.read())

        if not timesheets:
            st.error("No employee timesheets found in the PDF. Check the file format.")
        else:
            with st.spinner("Loading employee rates…"):
                rates = load_rates(excel_file.read())

            # Warn about missing employees
            missing = [ts["name"] for ts in timesheets if ts["name"].lower() not in rates]
            if missing:
                st.warning(
                    "⚠️ The following employees were **not found** in the rates file "
                    "(rate will be 0):\n\n" + "\n".join(f"- {n}" for n in missing)
                )

            with st.spinner("Generating Excel report…"):
                report_bytes = generate_report(timesheets, rates)

            st.success(f"✅ Done! {len(timesheets)} employee(s) processed.")

            # ── Preview summary ───────────────────────────────────────────
            st.subheader("Summary Preview")
            rows = []
            for ts in timesheets:
                r = rates.get(ts["name"].lower(), {})
                chargeable = sum(1 for d in ts["daily_data"] if d["Charged"] == "YES")
                travel_check = sum(1 for d in ts["daily_data"] if d["Charged"] == "CHECK")
                rem_rate = r.get("remuneration_rate", 0)
                cost_rate = r.get("cost_rate", 0)
                rows.append({
                    "Employee":            ts["name"],
                    "Project":             ts["project"],
                    "Month":               ts["month"],
                    "Chargeable Days":     chargeable,
                    "T Days (check)":      travel_check,
                    "Remuneration Total":  f"{chargeable * rem_rate:,.2f}",
                    "Cost Total":          f"{chargeable * cost_rate:,.2f}",
                })
            st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

            if travel_check > 0:
                st.info("ℹ️ Travel days (T) are marked **CHECK** in the detailed sheet — review manually.")

            # ── Download button ───────────────────────────────────────────
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            st.download_button(
                label="⬇️ Download Excel Report",
                data=report_bytes,
                file_name=f"Timesheet_Report_{timestamp}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
